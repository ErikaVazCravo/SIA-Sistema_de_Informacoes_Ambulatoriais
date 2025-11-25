# Essa conferência é feita a partir dos relatórios gerados 
# por prestador nas automações 07a e 07b
# Vlr. Bruto > EMISSÃO SÍNTESE DA PRODUCAO - BPA
# Produção BPA (BPAI e BPAC)

import os
import re
import openpyxl
from openpyxl.utils import get_column_letter

# ============== PARÂMETROS ==============
competencia = " 07/2025"  # mantenha no formato que aparece no TXT - coloca um espaço antes da competência
pasta = r"Coloque aqui o caminho da pasta onde estão os relatórios gerados"
relatorio = os.path.join(pasta, "--- Sintese da Producao BPA.xlsx")
# ========================================

def ajustar_largura_colunas(ws):
    dims = {}
    for row in ws.iter_rows(values_only=True):
        for i, value in enumerate(row, 1):
            v = "" if value is None else str(value)
            dims[i] = max(dims.get(i, 0), len(v))
    for i, width in dims.items():
        ws.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 10), 60)

def _parse_int(s):
    try:
        return int(s.strip())
    except:
        return 0

def processar_arquivo(caminho_arquivo, competencia, nome_estabelecimento):
    registros = []

    with open(caminho_arquivo, "r", encoding="utf-8", errors="ignore") as f:
        for linha in f:
            linha = linha.rstrip("\n\r")

            # Só linhas que começam com a competência EXATA (ex: " 07/2025")
            if not linha.startswith(competencia):
                continue

            cmp = linha[0:8].strip()
            flh = linha[8:12].strip()
            sq = linha[12:15].strip()
            proc = linha[15:27].strip()
            cbo = linha[27:34].strip()
            qt_prz = _parse_int(linha[34:42])
            vl_prz = linha[42:53].strip()
            qt_apvd = _parse_int(linha[53:61])
            vl_apvd = linha[61:72].strip()
            situacao = linha[72:150].strip()

            diferenca = qt_prz - qt_apvd

            registros.append([
                nome_estabelecimento,
                cmp, flh, sq, proc, cbo,
                qt_prz, vl_prz, qt_apvd, vl_apvd,
                diferenca,
                situacao
            ])
    return registros

# ------------------- Gerar Excel -------------------
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Síntese"
ws.append([
    "Estabelecimento", "Cmp", "Flh", "Sq", "Proc.", "CBO",
    "Qt.Prz.", "Vl.Prz.", "Qt.Apvd", "Vl.Apvd",
    "Qt.Prz - Qt.Apvd", "Situacao"
])

total_registros = 0
for nome_arquivo in os.listdir(pasta):
    caminho = os.path.join(pasta, nome_arquivo)
    if not os.path.isfile(caminho) or not nome_arquivo.lower().endswith(".txt"):
        continue

    # Extrair nome do estabelecimento a partir do nome do arquivo
    base = os.path.splitext(nome_arquivo)[0]  # remove .txt
    nome_estabelecimento = base.replace(" - PRODUCAO BPA", "").strip()

    registros = processar_arquivo(caminho, competencia, nome_estabelecimento)
    for reg in registros:
        ws.append(reg)
    total_registros += len(registros)

ajustar_largura_colunas(ws)
ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
ws.freeze_panes = "A2"

wb.save(relatorio)
print(f"Relatório gerado com {total_registros} registros em: {relatorio}")
