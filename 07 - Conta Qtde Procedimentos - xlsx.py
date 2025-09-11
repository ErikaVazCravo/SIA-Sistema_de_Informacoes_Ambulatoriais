# A partir dos relatórios emitidos na automação 06 - Relatório Analítico de procedimentos por unidade - físico.py,
# essa automação irá gerar um arquivo xlsx onde constará os procedimentos e seus totais de cada estabelecimento
# Iremos utilizar també o arquivo txt onde constam os números de CNES e Nomes dos Estabelecimentos (mesmo utilizado em 01 - Relatórios para prestadores.py)

import os
import re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# ===================== PARÂMETROS =====================
competencia = "202507"  # informe aqui a competência desejada (YYYYMM)
pasta = r"coloque aqui o caminho" # do diretório onde estão os relatórios gerados em 06 - Relatório Analítico de procedimentos por unidade - físico.py
arquivo_tabela = r"coloque aqui o caminho"  # Define o caminho do arquivo contendo os CNES e nomes dos estabelecimentos
relatorio = os.path.join(pasta, "--- Cont_proced.xlsx") # Nome do arquivo xlsx quwe será gerado
# =====================================================

def ajustar_largura_colunas(ws):
    """Ajusta largura aproximada das colunas pelo conteúdo."""
    dims = {}
    for row in ws.iter_rows(values_only=True):
        for i, value in enumerate(row, 1):
            v = "" if value is None else str(value)
            dims[i] = max(dims.get(i, 0), len(v))
    for i, width in dims.items():
        ws.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 10), 60)

def processar_arquivo(caminho_arquivo):
    """Lê o arquivo TXT e retorna contagens por procedimento e tipo (BPA/APAC)."""
    contagem = defaultdict(lambda: {"produzido": 0, "aprovado": 0})
    
    with open(caminho_arquivo, "r", encoding="utf-8", errors="ignore") as f:
        for linha in f:
            if not linha.startswith(competencia):
                continue

            procedimento = linha[7:17].strip()
            numero_apac = linha[18:31].strip()
            qtd_produzido = linha[39:50].strip()
            qtd_aprovado = linha[51:62].strip()

            tipo = "APAC" if numero_apac else "BPA"

            try:
                qtd_produzido = int(qtd_produzido)
            except:
                qtd_produzido = 0
            try:
                qtd_aprovado = int(qtd_aprovado)
            except:
                qtd_aprovado = 0

            if procedimento:
                chave = (tipo, procedimento)
                contagem[chave]["produzido"] += qtd_produzido
                contagem[chave]["aprovado"] += qtd_aprovado

    return contagem

# ------------------- 1) Ler nomes/códigos (NOMES_CNES) -------------------
codigo_para_nome = {}
with open(arquivo_tabela, "r", encoding="utf-8", errors="ignore") as file:
    for linha in file:
        m = re.search(r"(\d{3}\.\d{3}-\d|\d{14}|\d{7})", linha)
        if m:
            codigo = m.group()
            nome = linha[:m.start()].strip()
            codigo_limpo = re.sub(r"[.\-]", "", codigo)
            if len(codigo_limpo) > 7:
                codigo_limpo = codigo_limpo[-7:]  # usa os 7 últimos
            codigo_para_nome[codigo_limpo] = nome

# ------------------- 2) Renomear arquivos pela cauda de 7 dígitos -------------------
for nome_arquivo in os.listdir(pasta):
    caminho_antigo = os.path.join(pasta, nome_arquivo)
    if not os.path.isfile(caminho_antigo) or nome_arquivo == os.path.basename(relatorio):
        continue

    m = re.search(r"(\d{7})(?=\.[^.]+$)", nome_arquivo)
    if not m:
        continue
    codigo_arquivo = m.group(1)

    if codigo_arquivo in codigo_para_nome:
        novo_nome_base = codigo_para_nome[codigo_arquivo] + ".txt"
        caminho_novo = os.path.join(pasta, novo_nome_base)

        # evita sobrescrever
        if os.path.exists(caminho_novo):
            idx = 1
            base_sem_ext = os.path.splitext(novo_nome_base)[0]
            while True:
                tentativa = f"{base_sem_ext} ({idx}).txt"
                caminho_tentativa = os.path.join(pasta, tentativa)
                if not os.path.exists(caminho_tentativa):
                    caminho_novo = caminho_tentativa
                    break
                idx += 1

        os.rename(caminho_antigo, caminho_novo)
        print(f"Renomeado: {nome_arquivo} -> {os.path.basename(caminho_novo)}")

# ------------------- Gerar Excel -------------------
wb = Workbook()
ws = wb.active
ws.title = "Contagem Procedimentos"
ws.append(["Estabelecimento", "Tipo", "Procedimento", "Total Produzido", "Total Aprovado"])

for nome_arquivo in os.listdir(pasta):
    caminho = os.path.join(pasta, nome_arquivo)
    if not os.path.isfile(caminho) or not nome_arquivo.lower().endswith(".txt"):
        continue

    contagem = processar_arquivo(caminho)

    total_produzido = 0
    total_aprovado = 0

    for (tipo, proc), valores in sorted(contagem.items()):
        ws.append([nome_arquivo, tipo, proc, valores["produzido"], valores["aprovado"]])
        total_produzido += valores["produzido"]
        total_aprovado += valores["aprovado"]

    # Linha de total por estabelecimento
    ws.append([nome_arquivo, "TOTAL", "", total_produzido, total_aprovado])

# Ajustes finais
ajustar_largura_colunas(ws)
ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
ws.freeze_panes = "A2"

wb.save(relatorio)
print(f"Relatório gerado em: {relatorio}")
