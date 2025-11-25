# A partir dos relatórios emitidos no 06a (Relatório Analítico de Procedimentos por Unidade / Físico)

import os
import re
import openpyxl
from openpyxl.utils import get_column_letter

# ============== PARÂMETROS ==============
competencia = "202509"     # sem espaço — formato EXATO como aparece no TXT
pasta = r"coloque aqui o caminho da pasta onde estão os relatórios"
relatorio = os.path.join(pasta, "--- Analitico BPA APAC.xlsx")
# ========================================


def ajustar_largura_colunas(ws):
    dims = {}
    for row in ws.iter_rows(values_only=True):
        for i, value in enumerate(row, 1):
            v = "" if value is None else str(value)
            dims[i] = max(dims.get(i, 0), len(v))
    for i, width in dims.items():
        ws.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 10), 60)


def parse_linha(linha, competencia):
    """
    Extrai campos fixos da linha do relatório analítico.
    """
    if not linha.startswith(competencia):
        return None

    try:
        cmp = linha[0:6].strip()
        procedimento = linha[7:17].strip()
        apac = linha[18:32].strip()
        cbo = linha[32:38].strip()
        qtd_prod = linha[39:50].strip()
        qtd_apvd = linha[51:62].strip()
        situacao = linha[63:].strip()

        qtd_prod = int(qtd_prod) if qtd_prod.strip().isdigit() else 0
        qtd_apvd = int(qtd_apvd) if qtd_apvd.strip().isdigit() else 0

        tipo = "APAC" if apac else "BPA/RAAS"

        return {
            "cmp": cmp,
            "procedimento": procedimento,
            "apac": apac,
            "cbo": cbo,
            "qtd_prod": qtd_prod,
            "qtd_apvd": qtd_apvd,
            "situacao": situacao,
            "tipo": tipo
        }
    except:
        return None


def processar_arquivo(caminho_arquivo, competencia, nome_estabelecimento):
    # agrupa por (procedimento, tipo)
    agregados = {}

    with open(caminho_arquivo, "r", encoding="utf-8", errors="ignore") as f:
        for linha in f:
            linha = linha.rstrip("\n\r")
            dados = parse_linha(linha, competencia)
            if not dados:
                continue

            k = (dados["procedimento"], dados["tipo"])

            if k not in agregados:
                agregados[k] = {
                    "texto_proced": dados["procedimento"],
                    "tipo": dados["tipo"],
                    "qtd_prod": 0,
                    "qtd_apvd": 0,
                }

            agregados[k]["qtd_prod"] += dados["qtd_prod"]
            agregados[k]["qtd_apvd"] += dados["qtd_apvd"]

    # retorna lista de registros
    return [
        [
            nome_estabelecimento,
            v["tipo"],
            v["texto_proced"],
            v["qtd_prod"],
            v["qtd_apvd"]
        ]
        for v in agregados.values()
    ]


# ------------------- Gerar Excel -------------------
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Síntese"

ws.append([
    "Estabelecimento",
    "Tipo",
    "Procedimento",
    "Total Produzido",
    "Total Aprovado"
])

total_registros = 0

for nome_arquivo in os.listdir(pasta):
    caminho = os.path.join(pasta, nome_arquivo)
    if not os.path.isfile(caminho) or not nome_arquivo.lower().endswith(".txt"):
        continue

    # Nome = tudo antes do ".txt"
    nome_estabelecimento = os.path.splitext(nome_arquivo)[0].strip()

    registros = processar_arquivo(caminho, competencia, nome_estabelecimento)
    for reg in registros:
        ws.append(reg)
    total_registros += len(registros)

ajustar_largura_colunas(ws)
ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
ws.freeze_panes = "A2"

wb.save(relatorio)
print(f"Relatório gerado com {total_registros} registros em: {relatorio}")
