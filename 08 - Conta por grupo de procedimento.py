# Essa conferência é feita a partir dos relatórios
# Vlr. Bruto > EMISSÃO SÍNTESE DA PRODUCAO - BPA (R0507) e
# Vlr. Bruto > EMISSÃO SÍNTESE DA PRODUCAO - APAC (R0570)
# Resumo por grupo de procedimento (BPAI, BPAC e APAC) - Juntar 
# Qt.Prz. = inteiro | Vl.Prz. = decimal (2 casas, moeda/valor)

# ================================
# Resumo Sintético BPA + APAC
# ================================
# Renomear R0507 para BPA.txt
# Renomear R0570 para APAC.txt

import os
import openpyxl
from openpyxl.utils import get_column_letter
from collections import defaultdict
import re

# ============== PARÂMETROS ==============
pasta = r"coloque aqui o caminho da pasta onde estão os relatórios"
relatorio_excel = os.path.join(pasta, "--- Sintese_BPA_APAC.xlsx")
relatorio_txt = os.path.join(pasta, "--- Linhas_Consideradas.txt")
# ========================================

# ---------- Funções auxiliares ----------
def ajustar_largura_colunas(ws):
    dims = {}
    for row in ws.iter_rows(values_only=True):
        for i, value in enumerate(row, 1):
            v = "" if value is None else str(value)
            dims[i] = max(dims.get(i, 0), len(v))
    for i, width in dims.items():
        ws.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 10), 60)

def _parse_int(s):
    try:
        return int(s.strip())
    except:
        return 0

def _parse_float(s):
    try:
        return float(s.strip().replace(",", "."))
    except:
        return 0.0

def limpar_linha(linha):
    linha = linha.lstrip()
    linha = re.sub(r"[\x00-\x1F\x7F]+", "", linha)
    return linha

# ---------- Processadores ----------
def processar_bpa(caminho_arquivo, resumo, arquivo_log):
    padrao_comp = re.compile(r"^\d{2}/\d{4}")  # formato xx/xxxx
    with open(caminho_arquivo, "r", encoding="utf-8", errors="ignore") as f:
        for linha in f:
            linha_original = linha.rstrip("\n")
            linha = limpar_linha(linha)
            if not padrao_comp.match(linha):
                continue

            proc = linha[15:27].strip()
            qt_prz = _parse_int(linha[34:42])
            vl_prz = _parse_float(linha[42:53])

            grupo = proc[:2] if len(proc) >= 2 else "??"
            resumo[("BPA", grupo)]["qt_prz"] += qt_prz
            resumo[("BPA", grupo)]["vl_prz"] += vl_prz

            # pega a competência detectada
            competencia_detectada = linha[0:7]
            arquivo_log.write(f"[BPA {competencia_detectada}] {linha_original}\n")

def processar_apac(caminho_arquivo, resumo, arquivo_log):
    with open(caminho_arquivo, "r", encoding="utf-8", errors="ignore") as f:
        for linha in f:
            linha_original = linha.rstrip("\n")
            linha = limpar_linha(linha)
            if len(linha) < 82 or linha.strip().startswith("SQ "):
                continue
            if not (linha[0:2].isdigit() and linha[2] == " "):
                continue

            proc = linha[3:14].strip()
            qt_prz = _parse_int(linha[21:29])
            vl_prz = _parse_float(linha[52:64])

            if not proc:
                continue

            grupo = proc[:2] if len(proc) >= 2 else "??"
            resumo[("APAC", grupo)]["qt_prz"] += qt_prz
            resumo[("APAC", grupo)]["vl_prz"] += vl_prz

            # para APAC não há competência no início
            arquivo_log.write(f"[APAC] {linha_original}\n")

# ---------- Gerar Excel ----------
def gerar_excel(resumo, destino):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumo"

    ws.append(["Origem", "Grupo", "Qt.Prz.", "Vl.Prz."])

    for origem in ["BPA", "APAC"]:
        for g in [f"{i:02}" for i in range(1, 10)]:
            qt = resumo[(origem, g)]["qt_prz"]
            vl = resumo[(origem, g)]["vl_prz"]
            if qt == 0 and vl == 0:
                continue
            ws.append([origem, g, qt, vl])

        total_qt = sum(resumo[(origem, g)]["qt_prz"] for g in [f"{i:02}" for i in range(1, 10)])
        total_vl = sum(resumo[(origem, g)]["vl_prz"] for g in [f"{i:02}" for i in range(1, 10)])
        ws.append([origem, "TOTAL", total_qt, total_vl])

    ajustar_largura_colunas(ws)
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.freeze_panes = "A2"

    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = "0"
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.number_format = "#,##0.00"

    wb.save(destino)

# ================= EXECUÇÃO =================
resumo = defaultdict(lambda: {"qt_prz": 0, "vl_prz": 0.0})

with open(relatorio_txt, "w", encoding="utf-8") as log:
    for nome_arquivo in os.listdir(pasta):
        caminho = os.path.join(pasta, nome_arquivo)
        if not os.path.isfile(caminho):
            continue
        nome = nome_arquivo.lower()

        if nome.endswith("bpa.txt"):
            processar_bpa(caminho, resumo, log)
        elif nome.endswith("apac.txt"):
            processar_apac(caminho, resumo, log)

gerar_excel(resumo, relatorio_excel)

print(f"Resumo gerado em: {relatorio_excel}")
print(f"Linhas consideradas salvas em: {relatorio_txt}")
